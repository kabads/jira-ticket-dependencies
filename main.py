from jira import JIRA
import openpyxl
import os


def authenticate_and_get_ticket(jira_url, username, api_token, ticket_id):
    # Authenticate to JIRA
    jira = JIRA(server=jira_url, basic_auth=(username, api_token))
    
    # Get the ticket
    ticket = jira.issue(ticket_id)
    
    # Get dependencies (linked issues)
    dependencies = []

    # Get dependencies (linked issues)
    if 'issuelinks' in ticket.fields.__dict__:
        for link in ticket.fields.issuelinks:
            if hasattr(link, 'outwardIssue'):
                dependencies.append(link.outwardIssue.key)
            elif hasattr(link, 'inwardIssue'):
                dependencies.append(link.inwardIssue.key)

    # Get subtasks
    if 'subtasks' in ticket.fields.__dict__:
        for subtask in ticket.fields.subtasks:
            dependencies.append(subtask.key)
    return ticket, dependencies

def read_second_column_as_list(file_path):
    # Open the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the active sheet
    sheet = workbook.active
    # Read the second column
    second_column = [cell.value for cell in sheet['B'] if cell.value is not None]
    return second_column


def write_ticket_and_dependencies_to_excel(file_path, ticket_id, dependencies):
    # Open the workbook or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    # Select the active sheet or create one if it doesn't exist
    if 'Sheet' in workbook.sheetnames:
        sheet = workbook['Sheet']
    else:
        sheet = workbook.create_sheet('Sheet')
    # Find the next available row
    next_row = sheet.max_row + 1
    # Write the ticket_id in the first column
    sheet.cell(row=next_row, column=1, value=ticket_id)
    # Write the dependencies in the subsequent columns
    for col, dependency in enumerate(dependencies, start=2):
        sheet.cell(row=next_row, column=col, value=dependency)
    # Save the workbook
    workbook.save(file_path)

# Example usage
if __name__ == "__main__":
    output_file_path = 'jira_output.xlsx'
    tickets = read_second_column_as_list('jira.xlsx')
    print(tickets)
    jira_url = os.getenv('JIRA_URL')
    username = os.getenv('JIRA_USERNAME')
    api_token = os.getenv('JIRA_API_TOKEN')
    for ticket_id in tickets:
        print(f"Processing ticket: {ticket_id}")
        ticket, dependencies = authenticate_and_get_ticket(jira_url, username, api_token, ticket_id)
        # print(f"Ticket: {ticket}")
        # print(f"Dependencies: {dependencies}")
        write_ticket_and_dependencies_to_excel(output_file_path, ticket_id, dependencies)